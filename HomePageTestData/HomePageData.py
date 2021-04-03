import openpyxl

class HomePageTestData:

    #test_data = [{"firstname":"Havisha","lastname":"Vemuri","gender":"Female"},{"firstname":"Chaitanya","lastname":"Velamuri","gender":"Male"}]

    @staticmethod
    def testdata(colName):
        List = []
        Dict = {}
        xl = openpyxl.load_workbook("C:\\Users\\user\\Desktop\\TestDemo.xlsx")
        sheet = xl.active
        rows = sheet.max_row
        cols = sheet.max_column
        for row in range(1, rows + 1):
            if sheet.cell(row=row, column=1).value != colName:
                for col in range(2, cols + 1):
                    Dict[sheet.cell(row=1, column=col).value]=sheet.cell(row=row, column=col).value
                    #List.append(Dict)
                    print(Dict)
        return Dict